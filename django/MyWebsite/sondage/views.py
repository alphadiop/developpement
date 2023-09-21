"""
https://docs.djangoproject.com/fr/3.2/intro/tutorial04/ du 22-08-2021
http://localhost:8000/sondage/
http://localhost:8000/admin/

ModelName.objects.create(fieldname=value).save():create
ModelName.objects.all() : select
ModelName.objects.filter(fieldname=value): Filter
ModelName.objects.exclude(fieldname=value):Exclude
ModelName.objects.all().order_by('field_name') :Ordering
ModelName.objects.filter(fieldname=value).update(fieldname=value):Update
ModelName.objects.get(id=id).delete() : Delete
"""
#Username : admin
#Email address : alphadiop@gmail.com
#Password : diop
#MembreListView
import csv
import xlwt
import openpyxl
import pandas as pd
import io
from tablib import Dataset
from django.urls import reverse
from django.urls import reverse_lazy
from django.shortcuts import render, redirect
from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic.list import ListView
from django.views.generic import ListView
from django.views.generic.edit import FormView
#Personn
from sondage.resources import PersonResource


#Formulaires
from sondage.forms import PersonForm
from sondage.forms import QuestionForm
from sondage.forms import AccountForm
from sondage.forms import CsvImportForm
from sondage.forms import ContactForm
from sondage.forms import QuestionFormu

#import tables
from sondage.tables import MembreTable
from sondage.tables import PersonTable #pour afficher les tables
from sondage.tables import QuestionTable


#Modèle
from sondage.models import Question
from sondage.models import Account
from sondage.models import Member #toujour importer le modèle que l'on souhaite afficher
from sondage.models import Person
# Create your views here.

#get_name
#QuestionForm
class ContactView(FormView):
    template_name = 'sondage/contact_send_msg.html'
    form_class = ContactForm
    success_url = reverse_lazy('author-list')

    def form_valid(self, form):
        form.send_email()
        #return super().form_valid(form)
        return super(ContactView, self).form_valid(form)
        

#'sondage:account_list'
#https://www.ultimatedjango.com/learn-django/lessons/newedit-contact-enable-ajax-2/side/
#Créer un formulaire de modèle
class AccountList(ListView):
    model = Account
    #paginate_by = 12
    template_name = 'sondage/account_list.html' #tester 'sondage:account_list.html'
    context_object_name = 'accounts'
    #queryset = Account.objects.filter(title__icontains='war')[:5]
    
    def get_queryset1(self):
        account_list = Account.objects.filter(owner=self.request.user)
        return account_list
        
    def get_queryset(self):
        try:
            a = self.request.GET.get('account',)
        except KeyError:
            a = None
        if a:
            account_list = Account.objects.filter(
                name__icontains=a,
                owner=self.request.user
            )
        else:
            account_list = Account.objects.filter(owner=self.request.user)
        return account_list
        
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(AccountList, self).dispatch(*args, **kwargs)
        

      

@login_required()
def account_detail(request, uuid):
    account = Account.objects.get(uuid=uuid)
    if account.owner != request.user:
            return HttpResponseForbidden()
    contacts = Account.objects.filter(account=account)
    variables = {'account': account, 'contacts': contacts,}
    return render(request, 'sondage/account_detail.html', variables)

@login_required()
def account_cru(request, uuid=None):

    if uuid:
        account = get_object_or_404(Account, uuid=uuid)
        if account.owner != request.user:
            return HttpResponseForbidden()
    else:
        account = Account(owner=request.user)

    if request.POST:
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            redirect_url = reverse(
                'sondage.views.account_detail',
                args=(account.uuid,)
            )
            return HttpResponseRedirect(redirect_url)
    else:
        form = AccountForm(instance=account)

    variables = {
        'form': form,
        'account':account
    }

    template = 'sondage/account_cru.html'

    return render(request, template, variables)
   


   

# comment afficher le contenu d'une table
from sondage.models import Member #toujour importer le modèle que l'on souhaite afficher
from sondage.tables import MembreTable
from django.views.generic import ListView

#comment afficher une table sous django
#Notez que le modèle "Member" est défini dans le module "models.py"
#On créée mainteant le modèle à partir du modéle "Member"
class MembreListView(ListView):
    model = Member
    context_object_name = "membre"
    template_name = 'sondage/membre.html'
    #paginate_by = 2
    #queryset = Member.objects.filter(city="Guyancourt")
    
    
    
    
    
    
    

def register(request):
    if request.method == 'POST':
        print('hi hello')
        first_name = request.POST['first_name']
        email = request.POST['email']
        password = request.POST['password']
        user = User.objects.create_user(first_name=first_name, email=email, password=password)
        user.save()
        print('user Created')
        user.redirect('/')
    else:
        print('user not created')
        return render(request, 'register.html')
        

def accueil(request):
    return HttpResponse("Hello, world. You're at the sondage index.")
 #----------------------------------------------------------------------
 
 
 
def detail1(request, question_id):
    return HttpResponse("You're looking at question %s." % question_id)

from django.http import Http404
def detail2(request, question_id):
    try:
        question = Question.objects.get(pk=question_id)
    except Question.DoesNotExist:
        raise Http404("Question does not exist")
    return render(request, 'sondage/detail.html', {'question': question})



#Il est très courant d’utiliser get() et de lever une exception Http404 si l’objet n’existe pas. 
#Django fournit un raccourci. Voici la vue detail() 

def detail(request, question_id):
    question = get_object_or_404(Question, pk=question_id)
    return render(request, 'sondage/detail.html', {'question': question})
 

 #------------------------------------------------
 
def results1(request, question_id):
    response = "You're looking at the results of question %s."
    return HttpResponse(response % question_id)


from django.shortcuts import get_object_or_404, render
def results(request, question_id):
    question = get_object_or_404(Question, pk=question_id)
    return render(request, 'sondage/results.html', {'question': question})
    

from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse

from .models import Choice, Question

def vote(request, question_id):
    question = get_object_or_404(Question, pk=question_id)
    try:
        selected_choice = question.choice_set.get(pk=request.POST['choice'])
    except (KeyError, Choice.DoesNotExist):
        # Redisplay the question voting form.
        return render(request, 'sondage/detail.html', {
            'question': question,
            'error_message': "You didn't select a choice.",
        })
    else:
        selected_choice.votes += 1
        selected_choice.save()
        # Always return an HttpResponseRedirect after successfully dealing
        # with POST data. This prevents data from being posted twice if a
        # user hits the Back button.
        return HttpResponseRedirect(reverse('sondage:results', args=(question.id,)))

#request.POST est un objet similaire à un dictionnaire qui vous permet d’accéder aux données envoyées par leurs clés. 
#Dans ce cas, request.POST['choice'] renvoie l’ID du choix sélectionné, sous forme d’une chaîne de caractères. 
#Les valeurs dans request.POST sont toujours des chaînes de caractères.

#request.POST['choice'] lèvera une exception KeyError si choice n’est pas spécifié dans les données POST. 
#Le code ci-dessus vérifie qu’une exception KeyError n’est pas levée 
#et réaffiche le formulaire de question avec un message d’erreur si choice n’est pas rempli.


#Après l’incrémentation du nombre de votes du choix, le code renvoie une HttpResponseRedirect 
#plutôt qu’une HttpResponse normale. 
#HttpResponseRedirect prend un seul paramètre : 
#l’URL vers laquelle l’utilisateur va être redirigé (voir le point suivant pour la manière de construire cette URL dans ce cas).

#Comme le commentaire Python l’indique, vous devez systématiquement renvoyer 
#une HttpResponseRedirect après avoir correctement traité les données POST. 
#Ceci n’est pas valable uniquement avec Django, c’est une bonne pratique du développement Web.

#Dans cet exemple, nous utilisons la fonction reverse() dans le constructeur de HttpResponseRedirect. 
#Cette fonction nous évite de coder en dur une URL dans une vue. 
#On lui donne en paramètre la vue vers laquelle nous voulons rediriger 
#ainsi que la partie variable de l’URL qui pointe vers cette vue. 
#Dans ce cas, en utilisant l’URLconf défini dans la partie 3 de ce tutoriel, 
#l’appel de la fonction reverse() va renvoyer la chaîne de caractères :



def index1(request):
    latest_question_list = Question.objects.order_by('-pub_date')[:5]
    output = ', '.join([q.question_text for q in latest_question_list])
    return HttpResponse(output)

from django.template import loader
def index2(request):
    latest_question_list = Question.objects.order_by('-pub_date')[:5]
    template = loader.get_template('sondage/index.html')
    context = {
        'latest_question_list': latest_question_list,
    }
    return HttpResponse(template.render(context, request))
    
def index(request):
    latest_question_list = Question.objects.order_by('-pub_date')[:5]
    context = {'latest_question_list': latest_question_list}
    return render(request, 'sondage/index.html', context)
    
 #a fonction render() prend comme premier paramètre l’objet requête, un nom de gabarit comme deuxième paramètre 
 #et un dictionnaire comme troisième paramètre facultatif. 
 #Elle retourne un objet HttpResponse composé par le gabarit interprété avec le contexte donné
    

#render(request, template_name, context=None, content_type=None, status=None, using=None)
from django.contrib.auth.decorators import login_required
@login_required()
def affichage_data(request):
    if "GET" == request.method:
        return render(request, 'sondage/affichage_data.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        wb.get_sheet_names()[0] #le premier onglet
        worksheet = wb[wb.get_sheet_names()[0]]

        #getting a particular sheet by name out of many sheets
        #worksheet = wb["Sheet1"]
        #print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'sondage/affichage_data.html', {"excel_data":excel_data})    
   
from django.contrib import messages
#import_csv
#Import csv to modèle
def import_csv(request):
    if request.method == "POST":
        csv_file = request.FILES["csv_file"]
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a .csv file.')
        
        # Create Member objects from passed in data
        file_data = csv_file.read().decode('ISO-8859-1')          
        #lines = file_data.split("\n")
        io_string = io.StringIO(file_data)
        next(io_string)
        #line = file_data.split(";")
        #print(lines)
        for column in csv.reader(io_string, delimiter=';'):
            #print(column[1])
            created = Member.objects.update_or_create(
            #id=column[0], #ne pas prendre
            full_name=column[1], 
            email=column[2], 
            title=column[3], 
            city=column[4], 
        )
            return redirect(".")
    form = CsvImportForm()
    if form.is_valid():
        form.save()
    context = {"form": form}
    #print(Member.objects.all())
    return render(request, "admin/csv_form.html", context)
#sv_form   
#personn 


#fields = ["id","full_name", "email", "title","city"]
def import_csv_(request):
    form = CsvImportForm(request.POST)
    if form.is_valid():
        form.save()
        form = CsvImportForm()
        clients = Member.objects.get(activated=False)
        with open(cliente.file_name.path, 'r') as f:
            reader = csv.reader(f)

            for i, row in enumerate(reader):
                if i==0:
                    pass
                else:
                    row = "".join(row)
                    row = row.replace(";", " ")
                    row = row.split(" ")
                    #id = fields[0], 
                    full_name = row[1].capitalize()
                    email = row[2].capitalize()
                    title = row[3].capitalize()
                    city = row[4].capitalize()

                    value = Cliente.objects.create(
                        full_name=full_name,
                        email=email,
                        title=title,
                        city=city,
                    )

                    print('oggetto creato:', value.full_name, value.email, value.title, value.city)

            clients.activated = True
            clients.save()

    context = {'form': form}
    template = 'admin/csv_form.html'
    return render(request, template, context)

#Formulaires

#Module : forms.py 
#https://github.com/django-import-export/django-import-export/tree/main/import_export/templates/admin/import_export
#https://www.kaherecode.com/tutorial/developper-vos-applications-web-avec-django-partie-2
# Create your views here.

from sondage.form_membre import MemberForm #Faut toujour importer la classe "MemberForm" définie dans le moduele : "form_membre.py"
def homepage1(request):
    form = MemberForm() #le formulaire à afficher sur la page
    return render(request, 'sondage/member.html', locals())

def homepage2(request):
    if request.method == 'POST':
        # initialise le formulaire avec les données envoyées
        form = MemberForm(request.POST)

        # si les données sont valides
        if form.is_valid():
            form.save()  # enregistrer le membre en base de données
    else:
        form = MemberForm()  # le formulaire à afficher sur la page

    return render(request, 'sondage/member.html', locals())
 
def homepage3(request):
    if request.method == 'POST':
        # initialise le formulaire avec les données envoyées
        form = MemberForm(request.POST)

        # si les données sont valides
        if form.is_valid():
            form.save()  # enregistrer le membre en base de données

            # enregistrer un message de succès
            messages.success(request, 'Merci pour ton inscription!')
    else:
        form = MemberForm()  # le formulaire à afficher sur la page
    return render(request, 'sondage/member.html', locals())


def homepage(request):
    if request.method == 'POST':
        # initialise le formulaire avec les données envoyées
        form = MemberForm(request.POST)
        # si les données sont valides
        if form.is_valid():
            form.save()  # enregistrer le membre en base de données
            full_name = form.cleaned_data['full_name']
            email = form.cleaned_data['email']
            title = form.cleaned_data['title']
            city = form.cleaned_data['city']
            
            envoi = True
            form = MemberForm()  # créer un nouveau formulaire vide
            # enregistrer un message de succès
            messages.success(request, 'User saved successfully.')
        #else:
            #messages.error(request, 'The form is invalid.')
        #return render(request,'sondage/enregistrer_membre.html', {'form': form})
    else:
        form = MemberForm()  # le formulaire à afficher sur la page
    #return render(request,'sondage/contact.html', locals()) 
    #https://zestedesavoir.com/tutoriels/598/developpez-votre-site-web-avec-le-framework-django/263_premiers-pas/1524_les-formulaires/
    return render(request,'sondage/enregistrer_membre.html', {'form': form})

#https://www.djangoproject.com/start/
#---------------------------------------------------------------------------------------

    
    
def essai_view(request):
    """ Afficher tous les articles de notre blog """
    requete = Member.objects.all()
    table = MembreTable(requete) # Nous sélectionnons tous :
    #data = table.values() #data est une liste de dictionnaires
    #Entry.objects.filter(pub_date__year=2006)
    #Book.objects.all().aggregate(Avg('price'))
    return render(request, 'sondage/membre2.html', {'table':table}) # c'est le nom "datas" qui est utilisé dans le template


    
#Pagination de la table Html
#MembreTable du module tables.py
from django_tables2 import RequestConfig

def pagination(request):
    requete = Member.objects.all()
    data = requete.values() #data est une liste de dictionnaires
    table = MembreTable(data)
    table.paginate(page=request.GET.get('page', 1), per_page=25)
    #RequestConfig(request).configure(table)
    return render(request,'sondage/membre2.html', {'table': table})

#csv to model



from django.forms import modelformset_factory

def manage_authors(request):
    AuthorFormSet = modelformset_factory(Member, fields=('full_name', 'email','title', 'city'))
    if request.method == 'POST':
        formset = AuthorFormSet(request.POST, request.FILES)
        if formset.is_valid():
            formset.save()
            return redirect(listes)
    else:
        formset = AuthorFormSet(queryset=Member.objects.all())
    return render(request, 'sondage/manage_authors.html', {'formset': formset})
    
def listes(request):
    return render(request, 'sondage/manage_authors.html', locals())  
#------------------------------------------------------------------------------------

#alimentaer une table django


 #FromCsv.objects.create(username=csvusername, title=csvtitle, interests=csvinterests


def home_view(request): #ensuite faut mettre cette fontion dans le module url.py de l'application
    context ={}
    context['form']= QuestionForm()
    return render(request, "sondage/home.html", context)
    

def saisir_of_question(request):
    # if this is a POST request we need to process the form data
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = QuestionFormu(request.POST)
        # check whether it's valid:
        if form.is_valid():
            question_text = form.cleaned_data['question_text']
            pub_date = form.cleaned_data['pub_date']
            form.save() # enregistrer le membre en base de données
            form = QuestionFormu()  # créer un nouveau formulaire vide
            messages.success(request, 'Merci pour ton inscription!')
            # process the data in form.cleaned_data as required
            # ...
            # redirect to a new URL:
            #return HttpResponseRedirect(reverse('all-borrowed'))
            return HttpResponse('Hello from Django!')
    # if a GET (or any other method) we'll create a blank form
    else:
        form = QuestionFormu()
    return render(request, 'sondage/form.html',{'form': form})
 



 #alimenter la table Person via le formulaire
def alimenter_form_person(request):
    if request.method == 'POST':
        form = PersonForm(request.POST or None)
        if form.is_valid():
            form.save()
            form = PersonForm()
            return HttpResponse('Hello from Django!')
            #return redirect('.')
    else:
        form = PersonForm()
    return render(request, 'sondage/person_template_page.html', {'form':form})
    
 #https://docs.djangoproject.com/fr/3.2/topics/http/file-uploads/
 #pagination
 


#person_resource = PersonResource()
#dataset = person_resource.export()
#dataset.csv
#dataset.json
#dataset.yaml

#Filtrage des données
#person_resource = PersonResource()
#queryset = Person.objects.filter(location='Helsinki')
#dataset = person_resource.export(queryset)
#dataset.yaml


# def export_csv(request):
    # person_resource = PersonResource()
    # dataset = person_resource.export()
    # response = HttpResponse(dataset.csv, content_type='text/csv')
    # response['Content-Disposition'] = 'attachment; filename="persons.csv"'
    # return response
    
# def export_json(request):
    # person_resource = PersonResource()
    # dataset = person_resource.export()
    # response = HttpResponse(dataset.json, content_type='application/json')
    # response['Content-Disposition'] = 'attachment; filename="persons.json"'
    # return response
    
# def export_xls(request):
    # person_resource = PersonResource()
    # dataset = person_resource.export()
    # response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="persons.xls"'
    # return response
    
from datasets import load_dataset
import tablib
#Importation de données
def simple_upload(request):
    if request.method == 'POST':
        person_resource = PersonResource()
        dataset = Dataset()
        new_persons = request.FILES['myfile']

        imported_data = dataset.load(new_persons.read(),format='xlsx')
        result = person_resource.import_data(dataset, dry_run=True)  # Test the data impor
        #with open(new_persons, 'r') as fh:
            #imported_data = dataset.load(fh)
            #result = person_resource.import_data(dataset, dry_run=True)  # Test the data import

        if not result.has_errors():
            person_resource.import_data(dataset, dry_run=False)  # Actually import now
    return render(request, 'sondage/simple_upload.html')
    
    
    
#affi_form_person   
#afficher la table Person
#besoin du modèle : class Person
def afficher_table_person(request):
    requete = Person.objects.all()
    data = requete.values() #data est une liste de dictionnaires
    table = PersonTable(data)
    #table.paginate(page=request.GET.get('page', 1), per_page=25)
    return render(request,'sondage/membre2.html', {'table': table})

def afficher_table_question(request):
    requete = Person.objects.all()
    data = requete.values() #data est une liste de dictionnaires
    table = QuestionTable(data)
    #table.paginate(page=request.GET.get('page', 1), per_page=25)
    return render(request,'sondage/membre2.html', {'table': table})


                                



def export_users_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'

    writer = csv.writer(response)
    writer.writerow(['Username', 'First name', 'Last name', 'Email address'])

    users = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for user in users:
        writer.writerow(user)

    return response
    


def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Username', 'First name', 'Last name', 'Email address', ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response