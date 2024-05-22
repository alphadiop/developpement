from os import listdir, system, chdir
import os
#from this import d
import openpyxl
import datetime
from operator import itemgetter

"""wb = openpyxl.load_workbook("modele1.xlsx")
print(wb.sheetnames)
sheet = wb["Feuille1"]
cell = sheet["D1"]
cell = sheet.cell(4, 1)
# exemple code for x in range(2,7):
#                   cell = sheet.cell(x, 2)
#                   print(cell.value)
#sheet.max_column detemrine colonne max
#sheet.max_row    determine ligne max
print(cell.value)"""




###### Choix répértoire #######
def folderChoice():
  print("#################### CHOIX DU REPERTOIRE ###################")
  print
  print("Entrez le répertoire qui contient les fichiers xlsx (tableur excel au format xlsx)")
  workFolder = input("Veuillez entrez l'arboressence ex:  c:\mondossier :  ")
  return workFolder
  
#Conversion du dossier
def folderconvert(workFolder):
  currentFolder = ""
  for x in workFolder:
    if x == "\\":
      currentFolder = currentFolder + "/"
    else:
      currentFolder = currentFolder + x
  return currentFolder

#Message d'erreur
def erreur():
  os.system('cls')
  print()
  print("Il y a une erreur ce dossier ne contient pas que des fichiers xlsx ou contient des dossier également.")
  print("Veuillez mettre un repertoire ne contenant QUE les fichiers XLSX")
  print("ou les fichiers XLSX présent ne sont pas conforme.")
  print("Si vous avez déjà executé ce programme dans ce dossier il est probable que le fichier généré factureclients soit crée, il faut alors le supprimé.")
  input("Appuyez sur une touche pour terminer le programme.")
  exit()



def filetest():
  filelist = []
  for file in fileInFolderList:
    ext = file[-4:]
    if ext == "xlsx" :
      filelist.append(file)
    elif ext == "lsx#":
      continue
    else:
      erreur()
    #test si le fichier est bien un classeur ou pas
    
  print("Test de fichier terminé... Le dossier contient que les fichiers xlsx....")
  print()
  return filelist


def checkfile(checkfilelist):
  for file in checkfilelist:
    print("les fichiers testé sont : "+str(file))
    wb = openpyxl.load_workbook(file, data_only=True)
    print(wb.sheetnames)
    sheet = wb["Feuille1"]
    #cell = sheet.cell(4, 1)
    cell = sheet["D1"]
    if not cell.value == 'feuille de commande': erreur(); return
    cell = sheet["A7"]
    if not cell.value == 'clients': erreur(); return
    cell = sheet["B7"]
    if not cell.value == 'maxi': erreur(); return
    cell = sheet["C7"]
    if not cell.value == 'geant': erreur(); return
    cell = sheet["D7"]
    if not cell.value == 'miche': erreur(); return
    print("Les fichiers XLSX sont conforme.")
  return
    # exemple code for x in range(2,7):
    #                   cell = sheet.cell(x, 2)
    #                   print(cell.value)
    #sheet.max_column detemrine colonne max
    #sheet.max_row    determine ligne max
    #print(success)
    #cell = sheet["B4"]
    #date.append(cell.value)
    #datafilelist.append(file+" "+str(cell.value))


def takeclientlist(filelist):
  nombreclient = []
  for file in filelist:
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb["Feuille1"]
    print("\n"+file)
    x = 8
    while True:
      cell = sheet.cell(x, 1)
      if not cell.value:
        break
      else:
        print(cell.value)
        nombreclient.append(cell.value)
        x = x+1
  nombreclient = list(set(nombreclient)) #supprime les doublons
  return nombreclient
  

def getcustomersdata(allCustomersInFolderList, filelist):
  totalCustomersData = []
  customersData = []
  for customers in allCustomersInFolderList:
    customersData = []
    for file in filelist:
      wb = openpyxl.load_workbook(file, data_only=True)
      sheet = wb["Feuille1"]
      print("\n"+file)
      x = 7
      while True:
        cell = sheet.cell(x, 1)
        if not cell.value:
          break        
        else:
          if (cell.value) == customers:
            #print(customers+ " c'est le superclients "+cell.value)
            customersData.append([customers ,sheet["B4"].value, sheet["B"+str(x)].value, sheet["C"+str(x)].value, sheet["D"+str(x)].value, sheet["E"+str(x)].value, sheet["F"+str(x)].value, sheet["G"+str(x)].value, sheet["H"+str(x)].value])
            x = 7
            break
          else:
            x = x+1
    totalCustomersData.append(customersData)
  return totalCustomersData
  

currentfolderwin = str(folderChoice())
currentfolder = str(folderconvert(currentfolderwin))

#Liste tout les fichiers du repertoire de travail
fileInFolderList = (listdir(currentfolder))

#test des fichiers du repertoire du travail
filelist = filetest()
print("les fichiers retenu sont "+str(filelist))

os.chdir(currentfolderwin)
print(currentfolder)


checkfile(filelist)

allCustomersInFolderList = takeclientlist(filelist)

print(allCustomersInFolderList)



customersdatalist = getcustomersdata(allCustomersInFolderList, filelist)
print(customersdatalist)

for check in customersdatalist: #Transformer tout les NONE des cellules en 0 car dans la liste les cellules vides sont marqué NONE
  for x in check:
    i = 0
    for checkin in x:
      if checkin == None:
        print('bordel: '+str(checkin))
        print(x)
        print(i)
        x[i] = 0

      i += 1

print('initialisation terminer')

for toto in customersdatalist: #trier les données selon les clients
  get_n = itemgetter(1)
  myList = toto
  myList.sort(key=get_n)

"""for x in customersdatalist:
  for y in x:

    print(str(y)+"\n")"""
if not os.path.exists(str(currentfolder)+"/factureclients"):
  os.makedirs(str(currentfolder)+"/factureclients")
os.chdir(currentfolderwin+"/factureclients")
for data in customersdatalist:  #creation d'une feuille excel
  wbOutput = openpyxl.Workbook()
  sheet = wbOutput.active
  sheet["A1"] = "Facture"
  sheet["B6"] = "DATE"
  sheet["C6"] = "Maxi"
  sheet["D6"] = "Geant"
  sheet["E6"] = "Miche"
  sheet["F6"] = "Galette"
  sheet["G6"] = "Somun"
  sheet["H6"] = "Marguerite"
  sheet["I6"] = "Pide"

  linetable = 7
  for inDATA in data: #distribution des données
    filenameOutput = str(inDATA[0])
    date = inDATA[1].strftime("%d/%m/%Y")
    maxi = inDATA[2]
    geant = inDATA[3]
    miche = inDATA[4]
    galette = inDATA[5]
    somun = inDATA[6]
    marguerite = inDATA[7]
    pide = inDATA[8]
    sheet["A2"] = str(filenameOutput)
    sheet["B"+str(linetable)] = str(date)
    sheet["C"+str(linetable)] = int(maxi)
    sheet["D"+str(linetable)] = int(geant)
    sheet["E"+str(linetable)] = int(miche)
    sheet["F"+str(linetable)] = int(galette)
    sheet["G"+str(linetable)] = int(somun)
    sheet["H"+str(linetable)] = int(marguerite)
    sheet["I"+str(linetable)] = int(pide)
    linetable += 1

    



  
  wbOutput.save(str(filenameOutput+".xlsx"))